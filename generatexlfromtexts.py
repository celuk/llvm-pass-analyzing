import os
import pandas as pd


folder = "C:\\Users\\shc\\Desktop\\ubuntu"
xls = "C:\\Users\\shc\\Desktop\\yf.xlsx"

sub_folders = [name for name in os.listdir(folder) if(os.path.isdir)(os.path.join(folder,name))]

from openpyxl import load_workbook
wb2 = load_workbook(xls)

for i in sub_folders:
    wb2.create_sheet(i)
    
wb2.save(xls)

with pd.ExcelWriter(xls) as writer:
    for i in sub_folders:
        print(folder + "\\" + i + "\\" + "ac.txt")
        df = pd.read_csv(folder + "\\" + i + "\\" + "ac.txt", delim_whitespace=' ', error_bad_lines=False, engine='python', header=None, encoding='utf-8')
        df.to_excel(writer, sheet_name = i, engine='xlsxwriter')
